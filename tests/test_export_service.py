import openpyxl

from models.pickplace import PickPlaceData, PickPlaceComponent
from models.review import ReviewRecord
from services.export_service import ExportService


def _records():
    return [
        ReviewRecord(
            designator="C1", mpn="MPN1", layer="Top",
            old_x=1.0, old_y=2.0, old_rotation=0,
            new_x=3.0, new_y=4.0, new_rotation=90,
            status="Edited", remark="ok", review_time="2026-01-01 00:00:00",
        ),
        ReviewRecord(
            designator="C2", mpn="MPN2", layer="Bottom",
            old_x=5.0, old_y=6.0, old_rotation=0,
            status="Pending",
        ),
    ]


def test_export_report(tmp_path):
    path = tmp_path / "report.xlsx"
    ExportService.export_report(_records(), str(path))
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "Designator"
    assert ws.cell(row=2, column=1).value == "C1"
    assert ws.cell(row=2, column=7).value == 3.0  # New X (edited)
    assert ws.cell(row=3, column=7).value in (None, "")  # New X (pending -> empty)
    assert ws.cell(row=3, column=13).value == "Pending"
    wb.close()


def test_export_pickplace_fixed(tmp_path):
    data = PickPlaceData(
        headers=["Designator", "MPN", "Layer", "X", "Y", "Rotation"],
        components=[
            PickPlaceComponent(designator="C1", mpn="MPN1", layer="Top", x=1, y=2, rotation=0),
            PickPlaceComponent(designator="C2", mpn="MPN2", layer="Bottom", x=5, y=6, rotation=0),
        ],
        raw_data=[
            {"Designator": "C1", "MPN": "MPN1", "Layer": "Top", "X": 1, "Y": 2, "Rotation": 0},
            {"Designator": "C2", "MPN": "MPN2", "Layer": "Bottom", "X": 5, "Y": 6, "Rotation": 0},
        ],
    )
    path = tmp_path / "fixed.xlsx"
    ExportService.export_pickplace_fixed(_records(), data, str(path))
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    assert ws.cell(row=2, column=1).value == "C1"
    assert ws.cell(row=2, column=4).value == 3.0   # X overridden by new_x
    assert ws.cell(row=3, column=4).value == 5      # C2 unchanged
    wb.close()


def test_export_pickplace_fixed_skips_inactive(tmp_path):
    data = PickPlaceData(
        headers=["Designator", "X", "Y"],
        components=[
            PickPlaceComponent(designator="C1", x=1, y=2),
            PickPlaceComponent(designator="DELETED", x=3, y=4),
        ],
        raw_data=[
            {"Designator": "C1", "X": 1, "Y": 2},
            {"Designator": "DELETED", "X": 3, "Y": 4},
        ],
    )
    records = [
        ReviewRecord(designator="C1", old_x=1, old_y=2, status="OK",
                     new_x=1, new_y=2, new_rotation=0),
    ]
    path = tmp_path / "fixed2.xlsx"
    ExportService.export_pickplace_fixed(records, data, str(path))
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    assert ws.max_row == 2  # only C1 row
    wb.close()
