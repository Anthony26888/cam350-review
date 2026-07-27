import os
from typing import List

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from models.review import ReviewRecord
from models.pickplace import PickPlaceData


_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)


class ExportService:

    @staticmethod
    def export_report(records: List[ReviewRecord], file_path: str) -> str:
        os.makedirs(os.path.dirname(file_path), exist_ok=True)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Review Report"

        headers = [
            "Designator", "MPN", "Layer", "Old X", "Old Y", "Old Rotation",
            "New X", "New Y", "New Rotation",
            "Aligned X", "Aligned Y", "Aligned Rotation",
            "Status", "Remark", "Review Time",
        ]

        _write_header(sheet, headers)

        for row_idx, record in enumerate(records, start=2):
            sheet.cell(row=row_idx, column=1, value=record.designator)
            sheet.cell(row=row_idx, column=2, value=record.mpn)
            sheet.cell(row=row_idx, column=3, value=record.layer)
            sheet.cell(row=row_idx, column=4, value=record.old_x)
            sheet.cell(row=row_idx, column=5, value=record.old_y)
            sheet.cell(row=row_idx, column=6, value=record.old_rotation)
            is_edited = record.status in ("Edited", "OK")
            is_aligned = record.status in ("Aligned", "OK")
            sheet.cell(row=row_idx, column=7, value=record.new_x if is_edited else "")
            sheet.cell(row=row_idx, column=8, value=record.new_y if is_edited else "")
            sheet.cell(row=row_idx, column=9, value=record.new_rotation if is_edited else "")
            sheet.cell(row=row_idx, column=10, value=record.new_x if is_aligned else "")
            sheet.cell(row=row_idx, column=11, value=record.new_y if is_aligned else "")
            sheet.cell(row=row_idx, column=12, value=record.new_rotation if is_aligned else "")
            sheet.cell(row=row_idx, column=13, value=record.status)
            sheet.cell(row=row_idx, column=14, value=record.remark)
            sheet.cell(row=row_idx, column=15, value=record.review_time or "")

            _apply_row_style(sheet, row_idx, len(headers))

        _auto_width(sheet, headers)

        workbook.save(file_path)
        workbook.close()
        return file_path

    @staticmethod
    def export_pickplace_fixed(
        records: List[ReviewRecord],
        original_data: PickPlaceData,
        file_path: str,
    ) -> str:
        os.makedirs(os.path.dirname(file_path), exist_ok=True)

        modified = {r.designator: r for r in records if r.has_modifications}

        comp_by_des = {c.designator: c for c in original_data.components if c.designator}
        has_panel_instance = any(c.panel_instance is not None for c in original_data.components)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "PickPlace Fixed"

        headers = list(original_data.headers)
        if has_panel_instance and "Panel_Instance" not in [h.strip() for h in headers]:
            headers.append("Panel_Instance")
        _write_header(sheet, headers)

        panel_col_idx = None
        if has_panel_instance:
            for i, h in enumerate(headers):
                if h.strip() == "Panel_Instance":
                    panel_col_idx = i
                    break

        for row_idx, raw_row in enumerate(original_data.raw_data, start=2):
            des = str(raw_row.get("Designator", "")).strip()
            record = modified.get(des)
            comp = comp_by_des.get(des)

            for col_idx, header in enumerate(headers):
                if panel_col_idx is not None and col_idx == panel_col_idx:
                    value = comp.panel_instance if comp and comp.panel_instance is not None else ""
                    sheet.cell(row=row_idx, column=col_idx + 1, value=value)
                    continue

                value = raw_row.get(header, "")
                if record is not None:
                    hdr_lower = header.lower()
                    if hdr_lower == "x" and record.new_x is not None:
                        value = record.new_x
                    elif hdr_lower == "y" and record.new_y is not None:
                        value = record.new_y
                    elif hdr_lower == "rotation" and record.new_rotation is not None:
                        value = record.new_rotation
                sheet.cell(row=row_idx, column=col_idx + 1, value=value)

            _apply_row_style(sheet, row_idx, len(headers))

        _auto_width(sheet, headers)

        workbook.save(file_path)
        workbook.close()
        return file_path


def _write_header(sheet, headers: List[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER


def _apply_row_style(sheet, row_idx: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = sheet.cell(row=row_idx, column=col)
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(sheet, headers: List[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for row in sheet.iter_rows(
            min_col=col_idx, max_col=col_idx, min_row=2, max_row=sheet.max_row
        ):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        sheet.column_dimensions[chr(64 + col_idx)].width = min(max_len + 4, 30)
